# xlsx xls
#openpyxl -> xlsx
from openpyxl.reader.excel import load_workbook
def readXlsxFile(path):
    # 打开文件
    file = load_workbook(filename=path)
    #所有表格的名称(列表)
    #print(file.sheetnames)
    sheets = file.sheetnames
    #拿出一个表格
    sheet = file.worksheets[0]   #不是名称
    #最大行数
    #print(sheet.max_row)
    # 最大列数
    #print(sheet.max_column)
    # 表名
    #print(sheet.title)
    # 读取一张表的数据
    for lineNum in range(1,sheet.max_row+1):
        #print(lineNum)
        lineList = []
        for coluNum in range(1,sheet.max_column+1):
            value = sheet.cell(row=lineNum,column=coluNum).value
            # if value != None:
            lineList.append(value)
        print(lineList)

path = r"D:\box\excel\itemlist.xlsx"
readXlsxFile(path)
