# xlsx xls
#openpyxl -> xlsx
from openpyxl.reader.excel import load_workbook
def readXlsxFile(path):
    dic = {}
    # 打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    sheets = file.sheetnames
    print(sheets)
    for i in range(len(sheets)):
        # 拿出一个表格
        sheet = file.worksheets[i]
        #一张表所有数据
        sheetInfo = []
        # 读取一张表的数据
        for lineNum in range(1,sheet.max_row+1):
            lineList = []
            for coluNum in range(1,sheet.max_column+1):
                value = sheet.cell(row=lineNum,column=coluNum).value
                # if value != None:
                lineList.append(value)
            sheetInfo.append(lineList)
        #将一张表的数据存到字典中
        dic[sheets[i]] = sheetInfo
    return dic
path = r"D:\box\excel\itemlist.xlsx"
dic = readXlsxFile(path)
print(dic)
print(len(dic))
